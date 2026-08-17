"""
Load Section B AWS BoM unit prices live from the uploaded workbook.
Falls back to config.BOM values if the sheet cannot be parsed.
"""
from __future__ import annotations
import copy
import openpyxl
from .config import BOM as FALLBACK

# Map BoM serial number (col A) -> config key
_SERIAL_MAP = {
    1: "win_4v8g",   2: "win_4v16g",  3: "win_8v16g",  4: "win_8v32g",
    5: "win_16v32g", 6: "win_16v64g",
    7: "rhel_2v4g",  8: "rhel_2v8g",  9: "rhel_4v8g",  10: "rhel_4v16g",
    11:"rhel_4v32g", 12:"rhel_8v16g", 13:"rhel_8v32g",
    14:"rhel_16v32g",15:"rhel_16v64g",
    16:"ebs_128",    17:"ebs_256",    18:"ebs_512",
    19:"ebs_1024",   20:"ebs_2048",
    21:"s3_hot",
    29:"net_xfer",   30:"nat_data",   31:"static_ip",
    36:"app_lb",     37:"net_fw",     38:"waf",          39:"kms",
    40:"direct_connect",
}


def load(path_or_buffer) -> dict:
    """Return pricing dict {key: {unit, disc, basis, line, ...}}."""
    prices = copy.deepcopy(FALLBACK)
    try:
        wb = openpyxl.load_workbook(path_or_buffer, data_only=True)
        ws = wb["Section B AWS BoM"]
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            serial = row[0].value          # col A
            unit   = row[5].value          # col F — Indicative Per Unit Price
            disc   = row[6].value          # col G — Discount %
            if serial in _SERIAL_MAP and unit is not None and disc is not None:
                key = _SERIAL_MAP[serial]
                prices[key]["unit"] = float(unit)
                prices[key]["disc"] = float(disc) / 100.0
    except Exception:
        pass  # keep fallback silently
    return prices
