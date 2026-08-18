"""
DR / Cost Estimate sheet generator.
Input: AWS Pricing Calculator CSV export + Section B BoM xlsx.
Applies BoM pricing to BoM-covered services; 8% to rest.
EC2: family-level substitution with vCPU-ratio scaling (discount fixed).
EBS: slab pricing per BoM; snapshot separate at 8%.
"""
from __future__ import annotations
import os, re, math
import pandas as pd
import openpyxl
import copy
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from .bom    import load as load_bom
from .cur    import read_tabular
from .config import (
    BOM as BOM_FALLBACK, ebs_key, EC2_EXACT, EC2_ANCHOR, WIN_M5A_ANCHORS,
    ORANGE, GREEN, HDR_BLUE, COL_WIDTHS, NOTE_TEXT, NON_BOM_DISC,
)

_thin = Side(style="thin", color="BFBFBF")
_med  = Side(style="medium")
_bdr  = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

RHEL_DISC = 0.4128
WIN_DISC  = 0.4077

# Instance vCPU/mem resolution and family logic are shared with the billing
# engine so every tab applies identical rules (Case-2 fix + point 1).
from .pricing import instance_specs as _instance_specs, _base_family as _base_fam
from .config import WIN_M_FAMILIES


def _get_ec2_pricing(inst: str, vcpu: int, mem: int, os_cls: str, bom: dict):
    """Returns (bom_key or None, multiplier, anchor_desc or None)."""
    key = EC2_EXACT.get((os_cls, vcpu, mem))
    if key:
        return key, 1.0, None
    fam = _base_fam(inst)
    if os_cls == "rhel":
        if fam == "r6":
            ak, av = "rhel_4v32g", 4
            return ak, vcpu / av, f"r6a.xlarge anchor (4 vCPU, ₹{bom['rhel_4v32g']['unit']:,.2f})"
        if fam == "m6":
            ak, av = "rhel_16v64g", 16
            return ak, vcpu / av, f"m6g.4xlarge anchor (16 vCPU, ₹{bom['rhel_16v64g']['unit']:,.2f})"
    if os_cls == "win" and fam in WIN_M_FAMILIES:
        # Point 1: Windows m5a AND m6a scale through the m5a BoM lines.
        av = min((v for v in WIN_M5A_ANCHORS if v <= vcpu or v == 16), key=lambda x: abs(x - vcpu))
        ak = WIN_M5A_ANCHORS[av]
        return ak, vcpu / av, f"m5a anchor ({av} vCPU, ₹{bom[ak]['unit']:,.2f})"
    # Win r6 or unknown -> 8%
    return None, 1.0, None


def _read_raw_grid(path_or_buffer) -> pd.DataFrame:
    """
    Read a CSV or Excel file into a headerless, string-only DataFrame that is
    tolerant of ragged rows (AWS Calculator exports have short preamble rows
    followed by wider data rows, which trips pandas.read_csv). Excel goes
    through read_excel(header=None); CSV is parsed with the stdlib csv module
    and padded to a rectangle.
    """
    name = str(getattr(path_or_buffer, "name", path_or_buffer) or "").lower()
    is_excel = name.endswith((".xlsx", ".xls", ".xlsm"))
    if not is_excel and not name.endswith(".csv"):
        try:
            if hasattr(path_or_buffer, "read"):
                pos = path_or_buffer.tell(); head = path_or_buffer.read(2); path_or_buffer.seek(pos)
            else:
                with open(path_or_buffer, "rb") as fh: head = fh.read(2)
            is_excel = head[:2] == b"PK"
        except Exception:
            is_excel = False
    if is_excel:
        return pd.read_excel(path_or_buffer, header=None, dtype=str).fillna("")

    import csv as _csv, io as _io
    if hasattr(path_or_buffer, "read"):
        pos = path_or_buffer.tell()
        raw = path_or_buffer.read(); path_or_buffer.seek(pos)
        text = raw.decode("utf-8", "replace") if isinstance(raw, bytes) else raw
        fh = _io.StringIO(text); close = False
    else:
        fh = open(path_or_buffer, newline="", encoding="utf-8", errors="replace"); close = True
    try:
        rows = list(_csv.reader(fh))
    finally:
        if close:
            fh.close()
    maxc = max((len(r) for r in rows), default=1)
    rows = [r + [""] * (maxc - len(r)) for r in rows]
    return pd.DataFrame(rows).fillna("")


def _parse_csv(path_or_buffer):
    # AWS Pricing Calculator exports carry a variable number of preamble rows
    # before the real header (and Excel vs CSV exports differ), so instead of a
    # fixed skiprows we read raw (ragged-safe) and locate the header row.
    raw = _read_raw_grid(path_or_buffer)

    header_idx = None
    for i in range(min(40, len(raw))):
        vals = [str(x).strip() for x in raw.iloc[i].tolist()]
        if "Service" in vals:
            header_idx = i
            # Prefer a row that also has the config column (the true header).
            if any("Configuration" in v for v in vals):
                break

    if header_idx is None:
        raise ValueError(
            "This doesn't look like an AWS Pricing Calculator export — no "
            "'Service' column was found. Please upload the Calculator CSV/Excel "
            "export (the file with Service / Configuration summary / Monthly "
            "columns), not the BoM or a generated bill."
        )

    cols = [str(x).strip() for x in raw.iloc[header_idx].tolist()]
    df = raw.iloc[header_idx + 1:].copy()
    df.columns = cols
    df = df.reset_index(drop=True)
    # Drop fully-empty trailing rows.
    df = df[~(df == "").all(axis=1)].copy()

    for col in ("Monthly", "First 12 months total"):
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)
    # filter to group rows (when the Group hierarchy column is present)
    grp_col = next((c for c in df.columns if "Group hierarchy" in c), None)
    if grp_col:
        df = df[df[grp_col].astype(str).str.strip() != ""].copy()
    return df


def _parse_instance(cfg: str):
    os_ = "Windows" if "Windows Server" in cfg else ("RHEL" if "Red Hat" in cfg else "Linux")
    inst_m = re.search(r"EC2 instance \(([^)]+)\)", cfg)
    base_m = re.search(r"Baseline:\s*(\d+)", cfg)
    return os_, (inst_m.group(1) if inst_m else None), (int(base_m.group(1)) if base_m else 0)


def build_estimate(
    csv_path,
    bom_path,
    out_dir: str,
    title:   str = "DR Estimate",
    rate:    float = 91.0,
    date_label: str = "today",
) -> str:
    bom = load_bom(bom_path)
    df  = _parse_csv(csv_path)
    os.makedirs(out_dir, exist_ok=True)

    # Required columns from the Calculator export.
    if "Service" not in df.columns:
        raise ValueError(
            "The uploaded estimate file has no 'Service' column after parsing. "
            "Please upload the AWS Pricing Calculator CSV/Excel export."
        )
    if "Configuration summary" not in df.columns:
        # Some exports label it differently; try a close match, else blank it.
        alt = next((c for c in df.columns if "onfiguration" in c), None)
        df["Configuration summary"] = df[alt] if alt else ""

    svc = df["Service"].astype(str).str.strip()
    ec2 = df[svc == "Amazon EC2"].copy()
    ec2[["os_","inst","base"]] = ec2["Configuration summary"].apply(
        lambda c: pd.Series(_parse_instance(str(c))))
    non_ec2 = df[svc != "Amazon EC2"].copy()

    rows = []    # (sn, addl, cfg, sku, qty, I_formula, disc, is_bom, note_n)
    notes = []   # (n, text)

    sn = 1
    note_n = [0]

    def next_note():
        note_n[0] += 1
        return note_n[0]

    def bom_I(key, consumption=1.0, mult=1.0):
        b = bom[key]
        unit = b["unit"] * mult
        basis = b["basis"]
        if basis == 1:
            return f"=(({unit:.6f}*G{{r}}))/{rate}"
        return f"=((({unit:.6f}/{basis})*{round(consumption,4)})*G{{r}})/{rate}"

    # ── EC2 rows (grouped by OS then instance type) ──────────────────────
    for os_label in ["RHEL", "Windows"]:
        sub = ec2[ec2["os_"] == os_label].copy()
        if sub.empty: continue
        sub["_v"] = sub["inst"].map(lambda i: _instance_specs(str(i))[0])
        sub["_m"] = sub["inst"].map(lambda i: _instance_specs(str(i))[1])
        grp = sub.groupby("inst").agg(
            qty=("base","sum"), nrows=("inst","size"),
            apps=("Description", lambda s: "; ".join(sorted(set(str(x) for x in s)))),
            monthly=("Monthly","sum"),
            v=("_v","first"), m=("_m","first"),
        ).reset_index().sort_values(["v","m"])
        os_cls = "rhel" if os_label == "RHEL" else "win"
        os_name = "Red Hat Enterprise Linux" if os_label == "RHEL" else "Windows Server"

        for _, row in grp.iterrows():
            qty = max(int(row["qty"]), int(row["nrows"]))
            v, m = int(row["v"]), int(row["m"])
            inst = row["inst"]
            bkey, mult, anchor = _get_ec2_pricing(inst, v, m, os_cls, bom)
            disc = RHEL_DISC if os_cls == "rhel" else WIN_DISC
            nn = None
            if bkey:
                b_inst = bom[bkey].get("inst", bkey)
                nn = next_note()
                notes.append((nn,
                    f'Actual DR instance "{inst}" ({os_label}, {v} vCPU / {m} GB) priced using '
                    f'BoM "{b_inst}" (line {bom[bkey]["line"]}, ₹{bom[bkey]["unit"] * mult:,.2f})'
                    + (f" × {mult:.2f} vCPU ratio" if mult != 1.0 else " — exact match")
                    + f". Discount {disc:.2%} unchanged."
                ))
                I = bom_I(bkey, 1, mult)
            else:
                disc = NON_BOM_DISC
                I = str(round(float(row["monthly"]), 4))
            rows.append((sn, os_label, row["apps"], inst,
                         f"Chosen instance: {inst}  |  Operating system ({os_name})  |  vCPUs ({v})  |  Memory ({m} GiB)  |  Storage (EBS Only)",
                         qty, I, disc, bkey is not None, nn))
            sn += 1

    # ── Non-EC2 services ─────────────────────────────────────────────────
    for _, row in non_ec2.iterrows():
        svc = str(row["Service"]).strip()
        desc = str(row.get("Description","")).strip()
        cfg_sum = str(row.get("Configuration summary",""))
        monthly = float(row["Monthly"])

        # Point 3: drop zero-cost free-tier lines from the estimate too.
        if monthly <= 0:
            continue

        if "Direct Connect" in svc:
            # Point 2 — Direct Connect from BoM leased-line rate (all tabs).
            b = bom["direct_connect"]
            sp = re.search(r"(\d+(?:\.\d+)?)\s*(Gbps|G\b|Mbps|M\b)", cfg_sum, re.I)
            if sp:
                mbps = float(sp.group(1)) * (1000 if sp.group(2).lower().startswith("g") else 1)
            else:
                mbps = 1000.0
            qn_m = re.search(r"(?:quantity|connections?|ports?)\D*(\d+)", cfg_sum, re.I)
            qn = int(qn_m.group(1)) if qn_m else 1
            I = bom_I("direct_connect", mbps)
            rows.append((sn, "-", desc,
                         f"Dedicated DC-Cloud connect leased line | {mbps:.0f} Mbps × {qn}",
                         f"Direct Connect {mbps:.0f} Mbps leased line — BoM line 40 @ 75%",
                         qn, I, b["disc"], True, None))
            sn += 1
            continue

        if "Elastic Block Store" in svc or svc == "Amazon Elastic Block Store (EBS)":
            # Parse storage GB from config summary
            gb_m = re.search(r"Storage amount per volume \((\d+)\s*GB\)", cfg_sum)
            gb = int(gb_m.group(1)) if gb_m else 128
            vols_m = re.search(r"Number of volumes \((\d+)\)", cfg_sum)
            vols = int(vols_m.group(1)) if vols_m else 1
            key = ebs_key(gb)
            b = bom[key]
            I = bom_I(key, gb * vols)
            rows.append((sn, "-", desc, cfg_sum[:60],
                         f"EBS gp3 storage | {gb} GB × {vols} volumes | Priced per BoM Block SSD slab {gb} GB",
                         1, I, b["disc"], True, None))
        elif "Application Load Balancer" in svc or "Load Balancer" in svc:
            b = bom["app_lb"]
            I = bom_I("app_lb", 730)  # full month
            rows.append((sn, "-", desc, "Application Load Balancer | 1 TB/month",
                         f"Number of Application Load Balancers (1)", 1, I, b["disc"], True, None))
        elif "Network Firewall" in svc:
            b = bom["net_fw"]
            I = bom_I("net_fw", 1)
            rows.append((sn, "-", desc, "2 network firewall endpoints | 4Gbps | 1TB/month",
                         "Number of AWS Network Firewall endpoints (2), Usage per endpoint (730 hours), Data processed per month (1 TB)",
                         1, I, b["disc"], True, None))
        elif "WAF" in svc or "Web Application Firewall" in svc:
            req_m = re.search(r"([\d,]+)\s*million\s*requests", cfg_sum, re.I)
            req = int(req_m.group(1).replace(",","")) * 1_000_000 if req_m else 1_000_000
            b = bom["waf"]
            I = bom_I("waf", req)
            rows.append((sn, "-", desc, cfg_sum[:60],
                         f"Web ACL | {req:,} requests | Priced per BoM WAF",
                         1, I, b["disc"], True, None))
        else:
            rows.append((sn, "-", desc, cfg_sum[:60],
                         f"{svc} - not in BoM - standard pricing @ 8%",
                         1, str(round(monthly,4)), NON_BOM_DISC, False, None))
        sn += 1

    # ── Write workbook ───────────────────────────────────────────────────
    DR_HEADERS = ["Serial No.","OS","Application(s) / Description","Actual Instance",
                  "vCPU / Memory (Config)","AWS SKU","Quantity (X)","Monthly cost (USD)",
                  "Indicative Unit Cost","Discount %","Discounted Unit Cost",
                  "Monthly Cost Discounted (USD)","Converted Monthly Cost (INR)","Notes"]
    DR_WIDTHS  = [8, 9, 40, 15, 52, 52, 10, 14, 17, 10, 18, 22, 24, 8]

    orange_fill = PatternFill("solid", fgColor=ORANGE)
    green_fill  = PatternFill("solid", fgColor=GREEN)
    hdr_fill    = PatternFill("solid", fgColor=HDR_BLUE)
    no_fill     = PatternFill(fill_type=None)

    wb = openpyxl.Workbook(); ws = wb.active; ws.title = title[:28]
    for i, w in enumerate(DR_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    for c, h in enumerate(DR_HEADERS, 1):
        cell = ws.cell(1, c, h)
        cell.fill = hdr_fill
        cell.font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _bdr
    ws.row_dimensions[1].height = 42

    r = 2; first = r
    for sn_, os_, apps, inst, cfg, qty, I, disc, is_bom, nn in rows:
        Ival = I.replace("{r}", str(r)) if "{r}" in str(I) else str(I)
        ws.cell(r, 1,  sn_); ws.cell(r, 2,  os_)
        ws.cell(r, 3,  apps); ws.cell(r, 4, inst)
        ws.cell(r, 5,  cfg);  ws.cell(r, 6, cfg)  # cfg in both cols
        ws.cell(r, 7,  qty);  ws.cell(r, 8, Ival)
        ws.cell(r, 9,  f"=H{r}/G{r}"); ws.cell(r, 10, disc)
        ws.cell(r, 11, f"=I{r}*(100%-J{r})"); ws.cell(r, 12, f"=K{r}*G{r}")
        ws.cell(r, 13, f"=L{r}*{rate}")
        if nn:
            ws.cell(r, 14, nn).font = Font(name="Arial", size=8, vertAlign="superscript")
        fill = orange_fill if is_bom else no_fill
        for c_ in range(1, 14):
            cell = ws.cell(r, c_)
            cell.border = _bdr
            cell.font   = Font(name="Arial", size=10)
            cell.fill   = fill
            cell.alignment = Alignment(horizontal="center" if c_ in (1,2,7,10) else
                                       "right" if c_ in (8,9,11,12,13) else "left",
                                       vertical="center", wrap_text=True)
            if c_ == 10: cell.number_format = "0.00%"
            elif c_ in (8,9,11,12): cell.number_format = "$#,##0.00"
            elif c_ == 13: cell.number_format = "₹#,##0.00"
        ws.row_dimensions[r].height = 35
        r += 1
    last = r - 1

    # Total
    for c_ in range(1, 14):
        ws.cell(r, c_).border = Border(
            left=(_med if c_==1 else _thin), right=(_med if c_==13 else _thin),
            top=_thin, bottom=_med)
    ws.cell(r, 12, "Total").font = Font(name="Arial", size=11, bold=True)
    ws.cell(r, 12).fill = green_fill; ws.cell(r, 12).alignment = Alignment(horizontal="center", vertical="center")
    tot = ws.cell(r, 13, f"=SUM(M{first}:M{last})")
    tot.font = Font(name="Arial", size=11, bold=True); tot.fill = green_fill
    tot.number_format = "₹#,##0.00"; tot.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[r].height = 16; r += 1

    usd = ws.cell(r, 12, "Total (USD)")
    usd.font = Font(name="Arial", size=10, bold=True)
    usd.alignment = Alignment(horizontal="right", vertical="center")
    usd2 = ws.cell(r, 13, f"=SUM(L{first}:L{last})")
    usd2.font = Font(name="Arial", size=10, bold=True)
    usd2.number_format = "$#,##0.00"
    usd2.alignment = Alignment(horizontal="right", vertical="center")
    r += 1

    note_str = NOTE_TEXT.format(month=title)
    note_str += (" EC2: family-level BoM substitution with vCPU-ratio scaling; "
                 "discount % unchanged. Windows m-family (m5a/m6a) priced from "
                 "the m5a BoM lines; Windows r6a has no BoM equivalent → standard "
                 "8%. EBS slab pricing: ≤128→128, ≤256→256, ≤512→512, ≤1024→1024, "
                 ">1024→2048 GB slab. Direct Connect priced from BoM line 40 "
                 "(₹19,912.5 per 1 Gbps) @ 75%. Zero-cost free-tier lines excluded.")
    ws.cell(r, 1, note_str).font = Font(name="Arial", size=9, italic=True)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=13)
    ws.cell(r, 1).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[r].height = 60; r += 1

    ws.cell(r, 1, f"Converted Rate: Rs. {rate} as on {date_label}").font = Font(name="Arial", size=9, bold=True)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=13)
    ws.cell(r, 1).alignment = Alignment(horizontal="left", vertical="center")
    r += 2

    if notes:
        ws.cell(r, 1, "Working Notes").font = Font(name="Arial", size=11, bold=True); r += 1
        ws.cell(r, 1, "Note No.").font = Font(name="Arial", size=10, bold=True)
        ws.cell(r, 2, "Detailed Notes").font = Font(name="Arial", size=10, bold=True)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=13)
        ws.cell(r, 1).border = _bdr; ws.cell(r, 2).border = _bdr
        r += 1
        for nn, text in notes:
            ws.cell(r, 1, nn).alignment = Alignment(horizontal="center", vertical="top")
            ws.cell(r, 1).border = _bdr
            ws.cell(r, 2, text).font = Font(name="Arial", size=9)
            ws.cell(r, 2).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=13)
            ws.cell(r, 2).border = _bdr
            ws.row_dimensions[r].height = 60; r += 1

    ws.freeze_panes = "A2"
    out_path = os.path.join(out_dir, f"{title.replace(' ','_')}_BoM.xlsx")
    wb.save(out_path)
    return out_path
