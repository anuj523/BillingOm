"""
Render priced rows into the house-format .xlsx workbook.

House format (unchanged): orange BoM rows (F4B084), green Total (FF92D050),
blue header (4472C4), merged SN/Service cells, per-row formula chain
J=I/H, L=J·(1−K), M=L·H, N=M·rate, Section-D note, converted-rate footer,
and the numbered Working Notes table.

v3.2 presentation: a title block above the header, tighter column widths, and
print-ready page setup (landscape, fit-to-width, repeating header row, page
numbers) so the sheet no longer spills unreadably across many pages.
"""
from __future__ import annotations
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.properties import PageSetupProperties

from .config import COL_HEADERS, COL_WIDTHS, ORANGE, GREEN, HDR_BLUE, NOTE_TEXT

_thin  = Side(style="thin",   color="BFBFBF")
_med   = Side(style="medium")
_bdr   = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
_al_c  = Alignment(horizontal="center",  vertical="center", wrap_text=True)
_al_l  = Alignment(horizontal="left",    vertical="center", wrap_text=True)
_al_r  = Alignment(horizontal="right",   vertical="center")
_al_lt = Alignment(horizontal="left",    vertical="top",    wrap_text=True)

TITLE_TEXT = "IDBI Bank — AWS Monthly Billing"


def _font(name="Arial", size=10, bold=False, color="000000", italic=False):
    return Font(name=name, size=size, bold=bold, color=color, italic=italic)


def write(
    rows_tuples,          # [(sn, svc_name, is_bom, Row), ...]
    *,
    rate:          float,
    month_label:   str,
    date_label:    str,
    working_notes: list,  # [(n, text)]
    out_path:      str,
    sheet_title:   str,
    footer_note:   str | None = None,
    account_group: str | None = None,
) -> str:
    orange_fill = PatternFill("solid", fgColor=ORANGE)
    green_fill  = PatternFill("solid", fgColor=GREEN)
    hdr_fill    = PatternFill("solid", fgColor=HDR_BLUE)
    no_fill     = PatternFill(fill_type=None)

    # Notes column (last col) is present only when at least one note exists —
    # either a per-row note reference or a working-notes entry. Otherwise the
    # column is dropped entirely (presentation rule).
    has_notes = bool(working_notes) or any(
        getattr(row_obj, "note_n", None) for _, _, _, row_obj in rows_tuples
    )
    headers  = COL_HEADERS if has_notes else COL_HEADERS[:14]
    widths   = COL_WIDTHS  if has_notes else COL_WIDTHS[:14]
    ncols    = len(headers)
    last_col = get_column_letter(ncols)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_title
    ws.sheet_view.showGridLines = False

    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Title block ──────────────────────────────────────────────────────────
    ws.merge_cells(f"A1:{last_col}1")
    tcell = ws.cell(1, 1, TITLE_TEXT)
    tcell.font = _font("Arial", 15, bold=True, color="FFFFFF")
    tcell.alignment = Alignment(horizontal="center", vertical="center")
    for c in range(1, ncols + 1):
        ws.cell(1, c).fill = hdr_fill
    ws.row_dimensions[1].height = 30

    sub_bits = [f"Billing Month: {month_label}"]
    if account_group:
        sub_bits.append(f"Account Group: {account_group}")
    sub_bits.append(f"Prepared as on {date_label}")
    ws.merge_cells(f"A2:{last_col}2")
    scell = ws.cell(2, 1, "        |        ".join(sub_bits))
    scell.font = _font("Arial", 10, bold=True, color="203864")
    scell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 20

    ws.row_dimensions[3].height = 6   # thin spacer

    HDR_ROW = 4

    # ── Header row ───────────────────────────────────────────────────────────
    for c, h in enumerate(headers, 1):
        cell = ws.cell(HDR_ROW, c, h)
        cell.fill      = hdr_fill
        cell.font      = _font("Arial", 11, bold=True, color="FFFFFF")
        cell.alignment = _al_c
        cell.border    = _bdr
    ws.row_dimensions[HDR_ROW].height = 40

    r = HDR_ROW + 1
    first = r

    def apply_fill(row: int, is_bom: bool):
        fill = orange_fill if is_bom else no_fill
        for c in range(1, ncols + 1):
            ws.cell(row, c).fill = fill

    def style_row(row: int):
        for c in range(1, ncols + 1):
            cell = ws.cell(row, c)
            cell.border = _bdr
            cell.font   = _font()
            if c in (1, 2, 8, 11):
                cell.alignment = _al_c
            elif c in (9, 10, 12, 13, 14):
                cell.alignment = _al_r
            elif c == 15:
                cell.alignment = Alignment(horizontal="center", vertical="top")
            else:
                cell.alignment = _al_l
            if c == 11:
                cell.number_format = "0.00%"
            elif c in (9, 10, 12, 13):
                cell.number_format = "$#,##0.0000"
            elif c == 14:
                cell.number_format = "₹#,##0.00"

    # Track merged ranges
    sn_group_rows: dict[int, list] = {}

    for sn, svc_name, is_bom, row_obj in rows_tuples:
        sn_group_rows.setdefault(sn, []).append(r)

        ws.cell(r, 1, sn)
        ws.cell(r, 2, row_obj._aid)
        ws.cell(r, 3, row_obj._aname)
        ws.cell(r, 4, svc_name)
        ws.cell(r, 5, row_obj.additional)
        ws.cell(r, 6, row_obj.config)
        ws.cell(r, 7, row_obj.sku)
        ws.cell(r, 8, row_obj.qty)

        Ival = row_obj.i_formula.replace("{r}", str(r))
        ws.cell(r, 9,  Ival)
        ws.cell(r, 10, f"=I{r}/H{r}")
        ws.cell(r, 11, row_obj.discount)
        ws.cell(r, 12, f"=J{r}*(100%-K{r})")
        ws.cell(r, 13, f"=L{r}*H{r}")
        ws.cell(r, 14, f"=M{r}*{rate}")

        style_row(r)
        apply_fill(r, is_bom)

        if has_notes and row_obj.note_n:
            nc = ws.cell(r, 15, row_obj.note_n)
            nc.font = Font(name="Arial", size=8, vertAlign="superscript")

        r += 1

    last = r - 1

    # Merge SN and Service name for multi-account service groups
    for sn, row_list in sn_group_rows.items():
        if len(row_list) > 1:
            ws.merge_cells(start_row=row_list[0], start_column=1,
                           end_row=row_list[-1], end_column=1)
            ws.cell(row_list[0], 1).alignment = _al_c
            ws.merge_cells(start_row=row_list[0], start_column=4,
                           end_row=row_list[-1], end_column=4)
            ws.cell(row_list[0], 4).alignment = _al_l

    # ── Total row ────────────────────────────────────────────────────────────
    for c in range(1, ncols + 1):
        ws.cell(r, c).border = Border(
            left =(_med if c == 1     else _thin),
            right=(_med if c == ncols else _thin),
            top=_thin, bottom=_med,
        )
    ws.cell(r, 13, "Total").font = _font("Arial", 11, bold=True)
    ws.cell(r, 13).fill = green_fill
    ws.cell(r, 13).alignment = _al_c
    tot = ws.cell(r, 14, f"=SUM(N{first}:N{last})")
    tot.font = _font("Arial", 11, bold=True)
    tot.fill = green_fill
    tot.number_format = "₹#,##0.00"
    tot.alignment = _al_c
    ws.row_dimensions[r].height = 18
    total_row = r
    r += 2

    # ── Section-D note ───────────────────────────────────────────────────────
    note_text = footer_note or NOTE_TEXT.format(month=month_label)
    ws.cell(r, 1, note_text).font = _font("Arial", 9, italic=True)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
    ws.cell(r, 1).alignment = _al_l
    ws.row_dimensions[r].height = 42
    r += 1

    # ── Converted-rate footer ────────────────────────────────────────────────
    ws.cell(r, 1, f"Converted Rate: Rs. {rate} as on {date_label}").font = \
        _font("Arial", 9, bold=True)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
    ws.cell(r, 1).alignment = Alignment(horizontal="left", vertical="center")
    r += 2

    # ── Working notes ────────────────────────────────────────────────────────
    if working_notes:
        ws.cell(r, 1, "Working Notes").font = _font("Arial", 11, bold=True)
        r += 1
        ws.cell(r, 1, "Note No.").font = _font("Arial", 10, bold=True)
        ws.cell(r, 2, "Detailed Notes").font = _font("Arial", 10, bold=True)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=ncols)
        for c in (1, 2):
            ws.cell(r, c).border = _bdr
            ws.cell(r, c).fill = PatternFill("solid", fgColor="D9E1F2")
        r += 1
        for n, text in working_notes:
            ws.cell(r, 1, n).alignment = Alignment(horizontal="center", vertical="top")
            ws.cell(r, 1).border = _bdr
            ws.cell(r, 2, text).font = _font("Arial", 9)
            ws.cell(r, 2).alignment = _al_lt
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=ncols)
            ws.cell(r, 2).border = _bdr
            ws.row_dimensions[r].height = 54
            r += 1

    last_row = r - 1

    # ── Freeze panes: keep title + header visible while scrolling ─────────────
    ws.freeze_panes = f"A{first}"

    # ── Print / page setup ───────────────────────────────────────────────────
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize   = 9          # A4
    ws.page_setup.fitToWidth  = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.print_options.horizontalCentered = True
    ws.page_margins = PageMargins(left=0.3, right=0.3, top=0.5, bottom=0.5,
                                  header=0.2, footer=0.3)
    ws.print_title_rows = f"1:{HDR_ROW}"   # repeat title + header on every page
    ws.print_area = f"A1:{last_col}{last_row}"
    ws.oddFooter.right.text  = "Page &P of &N"
    ws.oddFooter.left.text   = TITLE_TEXT
    ws.oddFooter.left.size   = 8
    ws.oddFooter.right.size  = 8

    wb.save(out_path)
    return out_path
